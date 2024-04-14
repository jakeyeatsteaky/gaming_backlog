import pandas as pd
import openpyxl as xl
class Game:
    def __init__(self, title : str):
        self.game_title : str = title
        self.est_length : int = None
        self.excitement : int = None
        self.rolled_credits : bool = None
        self.reason: str = None
        self.time_spent : int = None
        self.over_the_course_of : int = None
        self.rating : int = None
        self.pre_thoughts: str = None
        self.post_thoughts: str = None
        



def print_game(game: Game):
    print(
        f"{game.game_title}:\n\t"
        f"{game.est_length if game.est_length is not None else 'N/A'} hours \n\t"
        f"{game.excitement if game.excitement is not None else 'N/A'} excitement \n\t"
        f"Rolled Credits: {game.rolled_credits if game.rolled_credits is not None else 'N/A'}\n\t"
        f"Because: {game.reason if game.reason is not None else 'N/A'}\n\t"
        f"Played for: {game.time_spent if game.time_spent is not None else 'N/A'}\n\t"
        f"Over the course of: {game.over_the_course_of if game.over_the_course_of is not None else 'N/A'} hours\n\t"
        f"I give it a: {game.rating if game.rating is not None else 'N/A'}\n\t"
        f"Thoughts: {game.pre_thoughts if game.pre_thoughts is not None else 'N/A'}\n\t"
        f"Thoughts after: {game.post_thoughts if game.post_thoughts is not None else 'N/A'}\n"
    )


def export_game_to_excel(game, filename='gaming_backlog.xlsx'):
    # Create a new workbook and select the active worksheet
    wb = xl.Workbook()
    ws = wb.active
    
    # Define the headers and game data as lists
    headers = ['Game Title', 'Estimated Length', 'Excitement', 'Rolled Credits', 
               'Reason', 'Time Spent', 'Over the Course Of', 'Rating', 
               'Pre-Thoughts', 'Post-Thoughts']
    data = [
        game.game_title, game.est_length, game.excitement, game.rolled_credits, 
        game.reason, game.time_spent, game.over_the_course_of, game.rating, 
        game.pre_thoughts, game.post_thoughts
    ]
    
    # Append the headers and the game data to the worksheet
    ws.append(headers)
    ws.append(data)
    
    # Formatting the header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = xl.styles.Font(bold=True)
        cell.alignment = xl.styles.Alignment(horizontal='center')
        
        # Auto-adjust column widths
        column_width = max(len(str(data[col - 1])), len(headers[col - 1])) + 2
        ws.column_dimensions[xl.utils.get_column_letter(col)].width = column_width
    
    # Save the workbook
    wb.save(filename)
    print(f'Data exported to {filename}')

def main():
    
    game = Game("Signalis")
    game.est_length = 10
    game.excitement = 3
    game.rolled_credits = False
    game.reason = "haven't started"
    print_game(game)

    export_game_to_excel(game)

if __name__ == "__main__":
    main()